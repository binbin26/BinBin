"""
Module xuất dữ liệu ra file Excel.
Hỗ trợ định dạng đẹp (kẻ bảng, tô màu header, tự động giãn cột).
"""

import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
import sys

# Import models
sys.path.append(str(Path(__file__).parent.parent))
from models.solution import Schedule

# Cấu hình logging
logger = logging.getLogger(__name__)

class Exporter:
    """
    Class chịu trách nhiệm xuất kết quả xếp lịch ra các định dạng file.
    """
    
    @staticmethod
    def export_to_excel(schedule: Schedule, file_path: str, proctors_dict: dict = None) -> bool:
        """
        Xuất lịch thi ra file Excel với định dạng đẹp.
        
        Args:
            schedule (Schedule): Đối tượng lịch thi cần xuất.
            file_path (str): Đường dẫn file lưu (.xlsx).
            proctors_dict (dict, optional): Dictionary map proctor_id -> Proctor object để lấy tên giám thị.
        
        Returns:
            bool: True nếu thành công.
        """
        try:
            if not schedule or not schedule.courses:
                logger.warning("Không có dữ liệu để xuất.")
                return False

            # 1. Chuẩn bị dữ liệu cho DataFrame
            if proctors_dict is None:
                proctors_dict = {}
            
            data = []
            for course in schedule.courses:
                # Lấy tên giám thị (hoặc ID nếu không tìm thấy)
                proctor_name = ""
                if course.assigned_proctor_id:
                    proctor_obj = proctors_dict.get(course.assigned_proctor_id)
                    if proctor_obj:
                        proctor_name = proctor_obj.name
                    else:
                        proctor_name = course.assigned_proctor_id  # Fallback: hiển thị ID
                
                # Sắp xếp lại thứ tự cột cho hợp lý
                data.append({
                    "Mã LHP": course.course_id,
                    "Tên học phần": course.name,
                    "Ngày thi": course.assigned_date,
                    "Giờ thi": course.assigned_time,
                    "Phòng thi": course.assigned_room,
                    "Giám thị": proctor_name,
                    "Địa điểm": course.location,
                    "Hình thức": course.exam_format,
                    "Sĩ số": course.student_count,
                    "Ghi chú": course.note
                })
            
            # Tạo DataFrame và sắp xếp theo Ngày -> Giờ -> Phòng
            df = pd.DataFrame(data)
            
            # Sắp xếp dữ liệu để file Excel dễ nhìn hơn
            if not df.empty and 'Ngày thi' in df.columns:
                df = df.sort_values(by=['Ngày thi', 'Giờ thi', 'Phòng thi'])

            # 2. Ghi ra file Excel dùng engine openpyxl
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Lich_Thi')
            
            # 3. Định dạng file Excel (Formatting)
            workbook = writer.book
            worksheet = writer.sheets['Lich_Thi']
            
            # --- Các kiểu định dạng ---
            
            # Font chữ và Căn lề Header
            header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") # Màu xanh dương
            header_align = Alignment(horizontal='center', vertical='center')
            
            # Font chữ và Căn lề Nội dung
            content_font = Font(name='Times New Roman', size=11)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Kẻ khung (Border)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

            # --- Áp dụng định dạng ---
            
            # Lấy kích thước bảng
            max_row = len(df) + 1
            max_col = len(df.columns)
            
            for col_idx, column_cells in enumerate(worksheet.columns, 1):
                # 1. Tự động giãn chiều rộng cột (Auto fit column width)
                length = max(len(str(cell.value)) for cell in column_cells)
                # Cộng thêm chút padding cho thoáng
                adjusted_width = (length + 4) * 1.2
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                
                for cell in column_cells:
                    # 2. Kẻ khung cho tất cả các ô
                    cell.border = thin_border
                    
                    # 3. Định dạng Header (Dòng 1)
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                    else:
                        # 4. Định dạng Nội dung (Các dòng sau)
                        cell.font = content_font
                        
                        # Căn trái cho Tên môn và Ghi chú, còn lại căn giữa
                        header_name = df.columns[col_idx - 1]
                        if header_name in ['Tên học phần', 'Ghi chú']:
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align

            # Lưu file
            writer.close()
            logger.info(f"Đã xuất file Excel thành công tại: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Lỗi khi xuất file Excel: {str(e)}")
            return False